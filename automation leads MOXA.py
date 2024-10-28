import pandas as pd
import os
from datetime import datetime
import numpy as np
from pathlib import Path
import openpyxl as op
from openpyxl.styles import Alignment, Font, Border, Side
import re
import time
import win32com.client as win32

folder_date = "28" # change
folder_month = "Oktober"
current_date = datetime.now().strftime("%Y%m%d")
dispatch_date = datetime.now().strftime("%d/%m/%Y")
filter_date = datetime.now().strftime("%d %B %Y")
column_date = ["Dispatch Date", "Update Status Date", "Tanggal Lahir"]
format_date = '%Y-%m-%d %H:%M:%S'

# path
base_path = Path(f"D:\\Daily MOXA\\blackup kirim dealer\\2024\\{folder_month}\\{folder_date}")
path_file = Path(f"D:\\Daily MOXA\\blackup kirim dealer\\2024\\{folder_month}\\{folder_date}\\DATA GABUNGAN LEADS FIFGROUP {current_date}.xlsx")
path = Path("D:\\Daily MOXA\\Master Leads Interest 2024.xlsx")
file = "D:\\Daily MOXA\\Automate Send to MD\\Email list.xlsx"
file_DaaS = "D:\Daily MOXA\DAAS\Rekap DAAS Februari 2023.xlsx"
path_folder = Path("D:\\Daily MOXA")
data_recap = Path("D:\\Daily MOXA\\backup\\Leads FIFGROUP Compile all MD.xlsx")
path_file = Path(f"D:\\Daily MOXA\\blackup kirim dealer\\2024\\{folder_month}\\{folder_date}\\DATA GABUNGAN LEADS FIFGROUP {current_date}.xlsx")
output_file_path_recap = os.path.join(path_folder, "Leads FIFGROUP Compile all MD.xlsx")

try:
    # data raw
    master = pd.read_excel(path, sheet_name="Oktober")
    email_list = pd.read_excel(file)
    dealer_DaaS = pd.read_excel(file_DaaS)
    
    # Filtering data master
    df_filtered = master[master["tgl"] == pd.to_datetime(filter_date)]# Double Check when daily task not sended on Time
    filter_DaaS = dealer_DaaS[dealer_DaaS['Dispatch Date'] == pd.to_datetime(filter_date)]
    
    # Set up Outlook
    outlook = win32.Dispatch("outlook.application")
except Exception as e:
    print(f"Failed to read master file: {e}")
    exit()

processed_dealers = set()

output_dir = f"D:\\Daily MOXA\\blackup kirim dealer\\2024\\{folder_month}\\{folder_date}"
os.makedirs(output_dir, exist_ok=True)

def adjust_column_width_and_format(filepath, *sheet_names, font_name='Calibri', font_size=11):
    workbook = op.load_workbook(filepath)
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        
        font_style = Font(name=font_name, size=font_size)
        alignment_style = Alignment(horizontal='left')  # Rata kiri

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for column_cells in sheet.columns:
            max_length = 0
            column = column_cells[0].column_letter
            
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                        cell.font = font_style
                        cell.alignment = alignment_style
                        cell.border = thin_border
                except:
                    pass

            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

    workbook.save(filepath)

def attach_files(mail, attachment_filenames):
    for filename, path_base in attachment_filenames:
        if path_base:
            attachment_path = path_base / filename
            if attachment_path.exists():
                mail.Attachments.Add(str(attachment_path))
                time.sleep(1)
            else:
                continue
        else:
            continue

def send_email(row, main_dealer_name, project_type, base_path): 
    if project_type == 'DaaS & MOXA':
        subject = f"Data leads FIFGROUP {filter_date} {row['Main Dealer']} (DaaS & MOXA)"
        attachment_filenames = [(f"Data leads FIFGROUP {current_date} {row['Main Dealer']}.xlsx", base_path),
                                (f"Data Leads FIFGROUP {current_date} {row['Main Dealer']} DaaS.xlsx", base_path)]
    elif project_type == 'DaaS':
        subject = f"Data leads FIFGROUP {filter_date} {row['Main Dealer']} (DaaS)"
        attachment_filenames = [(f"Data Leads FIFGROUP {current_date} {row['Main Dealer']} DaaS.xlsx", base_path)]
    elif project_type == 'MOXA':
        subject = f"Data leads FIFGROUP {filter_date} {row['Main Dealer']} (MOXA)"
        attachment_filenames = [(f"Data leads FIFGROUP {current_date} {row['Main Dealer']}.xlsx", base_path)]
    else:
        print(f"Invalid project type: {project_type}")
        return
    
    # create email
    mail = outlook.CreateItem(0)
    mail.To = row["to"]
    mail.CC = row["cc"]
    mail.Subject = subject
    mail.HTMLBody = f"""
<html>
<body style="font-family: Calibri, sans-serif; font-size: 11pt; color: black;">
<p>Dear Bapak & Ibu Yth,</p>

<p>Berikut terlampir data leads untuk pembiayaan motor baru dari aplikasi {project_type}.</p>

<p>Kami telah menambahkan waktu customer ingin dihubungi kembali, melalui channel apa customer ingin dihubungi kembali dan customer yang ingin melakukan pengajuan Syariah 
pada kolom remarks.</p>

<p>Terima kasih atas bantuan dan kerjasamanya,</p>

<p>Best Regards,<br>
Riyadh Akhdan Syafi<br>
<strong>CRM Data Mining</strong><br>
<a href="mailto:riyadh.asyafi@fifgroup.astra.co.id">riyadh.asyafi@fifgroup.astra.co.id</a>
</p>
</body>
</html>
"""
    attach_files(mail, attachment_filenames)

    # Display
    mail.Display()

    # Sending
    mail.Send()

def extract_dealer_name(filename):
    match = re.search(r'FIFGROUP \d+ (.+)\.xlsx', filename)
    if match :
        dealer_name = match.group(1)
        dealer_name = dealer_name.replace('DaaS', '').strip()
        return dealer_name
    return None

pemetaan_kolom = {
    "Id Leads Data User": "id",
    "Nama": "Nama",
    "Gender": "Gender",
    "Alamat": "Alamat",
    "Kelurahan": "Kelurahan",
    "Kecamatan": "Kecamatan",
    "Propinsi": "Propinsi",
    "Kota/Kabupaten": "Kota/Kabupaten",
    "No HP": "No HP",
    "MD (3 DIGIT)": "Main Dealer",
    "Pendidikan": "Pendidikan",
    "Tanggal Lahir": "Tanggal Lahir",
    "E-MAIL": "E-MAIL",
    "Dealer Sebelumnya (Jika Ada)": "Dealer Sebelumnya (Jika Ada)",
    "remarks": "Remarks/Keterangan"
}

kolom_akhir = [
    "id", "Nama", "Gender", "Alamat", "Kelurahan", "Kecamatan", "Kota/Kabupaten",
    "Propinsi", "No HP", "No Hp-2", "Sales Date", "Varian Motor", "Main Dealer",
    "Assign Dealer Code (5 DIGIT)", "Propensity", "Pekerjaan", "Pendidikan",
    "Pengeluaran", "Agama", "Tanggal Lahir", "Frame No Terakhir", "Jenis Penjualan",
    "Sales ID", "Nama Leasing Sebelumnya", "Nama salesman", "Source Leads",
    "Platform Data", "Dealer Sebelumnya (Jika Ada)", "Remarks/Keterangan",
    "Rekomendasi DP/Angsuran (Tenure)", "Varian motor yang diinginkan",
    "Warna varian motor", "E-MAIL", "FACEBOOK", "INSTAGRAM", "TWITTER"
]

try:
    df_pindah = df_filtered[list(pemetaan_kolom.keys())].rename(columns=pemetaan_kolom)
    df_pindah["No HP"] = df_pindah["No HP"].astype(str).apply(lambda x: "0" + x if not x.startswith("0") else x)
except:
    print("Failed to select and rename columns")

try:
    for kolom in kolom_akhir:
        if kolom not in df_pindah.columns:
            df_pindah[kolom] = np.nan
                
    df_pindah = df_pindah[kolom_akhir]
except:
    print("Failed to add missing columns")

# file path 
output_file_path = os.path.join(output_dir, f"DATA GABUNGAN LEADS FIFGROUP {current_date}.xlsx")

try:
    with pd.ExcelWriter(output_file_path, engine='xlsxwriter') as writer:
        df_pindah = df_pindah.fillna('')
        
        df_pindah.to_excel(writer, sheet_name='Sheet1',index=False)
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        border_format = workbook.add_format({'border': 1})
        date_format = workbook.add_format({'num_format': 'dd/mm/yyyy', 'border': 1})
        
        for row_num in range(len(df_pindah) + 1):
            for col_num, col in enumerate(df_pindah.columns):
                if row_num == 0:
                    value = col  # Header
                    worksheet.write(row_num, col_num, value, border_format)
                else:
                    value = df_pindah.iloc[row_num - 1, col_num]
                    
                    if col == 'Tanggal Lahir':
                        worksheet.write(row_num, col_num, value, date_format)
                    else:
                        worksheet.write(row_num, col_num, value, border_format)
        
        for idx, col in enumerate(df_pindah.columns):
            max_len = max(df_pindah[col].astype(str).map(len).max(), len(col)) + 2
            worksheet.set_column(idx, idx, max_len)
            
    print(f"File disimpan dengan format yang sama di {output_file_path}")
except Exception as e:
    print(f"Failed to save the output file: {e}")

file = f"D:\\Daily MOXA\\blackup kirim dealer\\2024\\{folder_month}\\{folder_date}\\DATA GABUNGAN LEADS FIFGROUP {current_date}.xlsx"

try:
    extracted_file = pd.read_excel(file)
except Exception as e:
    print(f"Failed to read extracted file: {e}")
    exit()

unique_values = extracted_file["Main Dealer"].unique()

for unique in unique_values:
    if unique in ["PT MPM - MALANG", "PT MPM - SURABAYA"]:
        kolom_mpm = ["id", "Nama", "No HP", "Kota/Kabupaten", "Kecamatan", "Alamat", "Main Dealer"]
        df_final = extracted_file[extracted_file["Main Dealer"] == unique].copy()
        df_final = df_final[kolom_mpm]
        df_final["No HP"] = df_final["No HP"].astype(str).apply(lambda x: "0" + x if not x.startswith("0") else x)

        output_path = os.path.join(output_dir, f"Data Leads FIFGROUP {current_date} {unique}.xlsx")
        try:
            with pd.ExcelWriter(output_path, engine="xlsxwriter") as writer:
                df_final.to_excel(writer, sheet_name="Sheet1", index=False)
                worksheet = writer.sheets["Sheet1"]
                
                for idx, col in enumerate(df_final.columns):
                    max_len = max(df_final[col].astype(str).map(len).max(), len(col)) + 2
                    worksheet.set_column(idx, idx, max_len)
            print(f"File has been Splitting {unique}")
        except Exception as e:
            print(f"Failed to save the file for {unique}: {e}")
    else:
        df_output = extracted_file[extracted_file["Main Dealer"] == unique].copy()
        df_output["No HP"] = df_output["No HP"].astype(str).apply(lambda x: "0" + x if not x.startswith("0") else x)

        output_path = os.path.join(output_dir, f"Data Leads FIFGROUP {current_date} {unique}.xlsx")
        try:
            with pd.ExcelWriter(output_path, engine="xlsxwriter") as writer:
                df_output.to_excel(writer, sheet_name="Sheet1", index=False)
                worksheet = writer.sheets["Sheet1"]
                
                for idx, col in enumerate(df_output.columns):
                    max_len = max(df_output[col].astype(str).map(len).max(), len(col)) + 2
                    worksheet.set_column(idx, idx, max_len)
            print(f"File has been Splitting {unique}")
        except Exception as e:
            print(f"Failed to save the file for {unique}: {e}")

# filtering data
# Moxa
try:
    dealer = pd.read_excel(path_file)
except Exception as e:
    print(f"reading file has been error {e}")

main_dealer_filtered = dealer['Main Dealer'].unique()
filter_email = email_list[email_list['Main Dealer'].isin(main_dealer_filtered)]
filter_email_MD = filter_email['Main Dealer'].unique()

# DaaS
DaaS_main_dealer = filter_DaaS['Main Dealer'].unique()
filter_email_DaaS = email_list[email_list['Main Dealer'].isin(DaaS_main_dealer)]
filter_email_DaaS_MD = filter_email_DaaS['Main Dealer'].unique()

# Moxa DaaS
over_lapping_maindealer = set(dealer['Main Dealer']).intersection(filter_DaaS["Main Dealer"])
over_lapping_maindealer_list = list(over_lapping_maindealer)
filter_email_DaaS_MOXA = email_list[email_list['Main Dealer'].isin(over_lapping_maindealer)]

try:
    with os.scandir(base_path) as entries:
        for entry in entries:
            if entry.is_file() and entry.name.endswith('.xlsx'): 
                dealer_name = extract_dealer_name(entry.name)

                if dealer_name is None or dealer_name in processed_dealers:
                    continue

                if dealer_name in over_lapping_maindealer_list:
                    # Send DaaS & MOXA email
                    for index, row in filter_email_DaaS_MOXA.iterrows():
                        if dealer_name == row['Main Dealer']:
                            send_email(row, dealer_name, 'DaaS & MOXA', base_path)
                            processed_dealers.add(dealer_name)
                            print(f"Sending DaaS & MOXA email to {dealer_name}")
                        else:
                            continue

                elif dealer_name not in over_lapping_maindealer_list:
                    if dealer_name in filter_email_DaaS_MD:
                        for index, row in filter_email_DaaS.iterrows():
                            if dealer_name == row['Main Dealer']:
                                send_email(row, dealer_name, 'DaaS', base_path)
                                processed_dealers.add(dealer_name)
                                print(f"Sending DaaS email to {dealer_name}")
                            else:
                                continue
                    
                    elif dealer_name in filter_email_MD:
                        for index, row in filter_email.iterrows():
                            if dealer_name == row['Main Dealer']:
                                send_email(row, dealer_name, 'MOXA', base_path)
                                processed_dealers.add(dealer_name)
                                print(f"Sending MOXA email to {dealer_name}")
                            else:
                                continue
except Exception as e :
    print(f"error occurs {e}")
    
try:
    df_recap = pd.read_excel(data_recap)
except Exception as e:
    print("data error : {e}")

# Membaca file Excel dan menggabungkannya
try:
    df_recap = pd.read_excel(data_recap)
    df_daily = pd.read_excel(path_file)
    df_daily['Source Leads'] = 'FIF'
    df_daily['Platform Data'] = 'MOXA'
    id_daily = df_daily['id'].to_list()
    df_merge = pd.concat([df_recap, df_daily], axis=0)
    try:
        for column in column_date:
            df_merge[column] = pd.to_datetime(df_merge[column], format='%Y-%m-%d %H:%M:%S', errors='coerce')
            df_merge[column].fillna(pd.to_datetime(df_merge[column], format='%d/%m/%Y', errors='coerce'), inplace=True)
            df_merge[column].fillna(pd.to_datetime(df_merge[column], format='%m/%d/%Y', errors='coerce'), inplace=True)
            df_merge[column] = df_merge[column].dt.strftime("%d/%m/%Y")
        df_merge['No HP'] = df_merge['No HP'].astype(str).apply(lambda x: "0" + x if not x.startswith("0") else x)
        df_merge.loc[df_merge['id'].isin(id_daily), "Dispatch Date"] = dispatch_date
    except Exception as e:
        print(f" error while running df_merge : {e}")
except Exception as e:
    print(f"error while running: {e}")

# Menulis data ke Excel menggunakan pandas dan XlsxWriter
try:
    with pd.ExcelWriter(output_file_path_recap, engine='xlsxwriter') as writer:
        df_merge.to_excel(writer, sheet_name="concate", index=False)
except Exception as e:
    print(f"Error detail: {e}")

# Setelah file Excel dibuat, buka file tersebut dengan openpyxl untuk menambahkan border dan mengatur lebar kolom
wb = op.load_workbook(output_file_path_recap)
ws = wb.active  # Menggunakan sheet pertama

# Menentukan style border
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

# Menyusun lebar kolom otomatis dan menambahkan border
for col in range(1, 46):  # Kolom A sampai AT (sesuai jumlah kolom yang Anda inginkan)
    max_length = 0
    col_letter = ws.cell(row=1, column=col).column_letter
    for row in range(1, ws.max_row + 1):  # Iterasi semua baris
        cell = ws.cell(row=row, column=col)
        # Menambahkan border
        cell.border = thin_border
        # Mengatur lebar kolom otomatis
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))

    # Mengatur lebar kolom berdasarkan panjang maksimum
    adjusted_width = max_length + 2
    ws.column_dimensions[col_letter].width = adjusted_width

hidden_column = ['C', 'D', 'E', 'F', 'G', 'H', 'J', 'K', 'L', 'N', 'O', 'P', 'Q', 'R', 'S', 
                   'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 
                   'AH', 'AI', 'AJ']

for hidden in hidden_column:
    ws.column_dimensions[hidden].hidden = True

# Menyimpan file yang sudah diubah
# output_modified_file_path = os.path.join(path_folder, "testing_concate.xlsx")
wb.save(output_file_path_recap)

print(f"File berhasil disimpan di: {output_file_path_recap}")