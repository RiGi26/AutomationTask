from datetime import datetime
import pandas as pd
from pathlib import Path
import os
import openpyxl as op
from openpyxl.styles import Alignment, Font, Border, Side

folder_date = "25"  # change
folder_month = "Oktober"
current_date = datetime.now().strftime("%Y%m%d")
dispatch_date = datetime.now().strftime("%d/%m/%Y")
column_date = ["Dispatch Date", "Update Status Date", "Tanggal Lahir"]
format_date = '%Y-%m-%d %H:%M:%S'

path_folder = Path("D:\\Daily MOXA")
data_recap = Path("D:\\Daily MOXA\\backup\\Leads FIFGROUP Compile all MD.xlsx")
path_file = Path(f"D:\\Daily MOXA\\blackup kirim dealer\\2024\\{folder_month}\\{folder_date}\\DATA GABUNGAN LEADS FIFGROUP {current_date}.xlsx")
output_file_path = os.path.join(path_folder, "Leads FIFGROUP Compile all MD.xlsx")

df_recap = pd.read_excel(data_recap)

# Membaca file Excel dan menggabungkannya
try:
    df_recap = pd.read_excel(data_recap)
    df_daily = pd.read_excel(path_file)
    df_daily['Source Leads'] = 'FIF'
    df_daily['Platform Data'] = 'MOXA'
    id_daily = df_daily['id'].to_list()
    df_merge = pd.concat([df_recap, df_daily], axis=0)
    try:
        for column in column_date:
            df_merge[column] = pd.to_datetime(df_merge[column], format='%Y-%m-%d %H:%M:%S', errors='coerce')
            df_merge[column].fillna(pd.to_datetime(df_merge[column], format='%d/%m/%Y', errors='coerce'), inplace=True)
            df_merge[column].fillna(pd.to_datetime(df_merge[column], format='%m/%d/%Y', errors='coerce'), inplace=True)
            df_merge[column] = df_merge[column].dt.strftime("%d/%m/%Y")
        df_merge['No HP'] = df_merge['No HP'].astype(str).apply(lambda x: "0" + x if not x.startswith("0") else x)
        df_merge.loc[df_merge['id'].isin(id_daily), "Dispatch Date"] = dispatch_date
    except Exception as e:
        print(f" error while running df_merge : {e}")
except Exception as e:
    print(f"error while running: {e}")

# Menulis data ke Excel menggunakan pandas dan XlsxWriter
try:
    with pd.ExcelWriter(output_file_path, engine='xlsxwriter') as writer:
        df_merge.to_excel(writer, sheet_name="concate", index=False)
except Exception as e:
    print(f"Error detail: {e}")

# Setelah file Excel dibuat, buka file tersebut dengan openpyxl untuk menambahkan border dan mengatur lebar kolom
wb = op.load_workbook(output_file_path)
ws = wb.active  # Menggunakan sheet pertama

# Menentukan style border
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

# Menyusun lebar kolom otomatis dan menambahkan border
for col in range(1, 46):  # Kolom A sampai AT (sesuai jumlah kolom yang Anda inginkan)
    max_length = 0
    col_letter = ws.cell(row=1, column=col).column_letter
    for row in range(1, ws.max_row + 1):  # Iterasi semua baris
        cell = ws.cell(row=row, column=col)
        # Menambahkan border
        cell.border = thin_border
        # Mengatur lebar kolom otomatis
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))

    # Mengatur lebar kolom berdasarkan panjang maksimum
    adjusted_width = max_length + 2
    ws.column_dimensions[col_letter].width = adjusted_width

hidden_column = ['C', 'D', 'E', 'F', 'G', 'H', 'J', 'K', 'L', 'N', 'O', 'P', 'Q', 'R', 'S', 
                   'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 
                   'AH', 'AI', 'AJ']

for hidden in hidden_column:
    ws.column_dimensions[hidden].hidden = True

# Menyimpan file yang sudah diubah
# output_modified_file_path = os.path.join(path_folder, "testing_concate.xlsx")
wb.save(output_file_path)

print(f"File berhasil disimpan di: {output_file_path}")